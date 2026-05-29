import io, os, uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import (Blueprint, render_template, redirect, url_for,
                   request, flash, send_file, current_app)
from flask_login import login_required, current_user

from extensions import db
from models.team import Team, Player, TeamPost

from utils.permissions import require_manager

team_bp = Blueprint('team', __name__)

POSITIONS_FUTSAL = ['GR', 'FIX', 'ALA-D', 'ALA-E', 'PV']
POSITIONS_FOOTBALL = ['GR', 'DD', 'DC', 'DE', 'MDC', 'MC', 'MO', 'ED', 'EE', 'PL', 'CA']
ALL_POSITIONS = sorted(set(POSITIONS_FUTSAL + POSITIONS_FOOTBALL))
FORMATIONS_FUTSAL = ['1-2-1', '1-3', '2-2', '3-1', '4-0']
FORMATIONS_FOOTBALL = ['4-3-3', '4-4-2', '4-2-3-1', '3-5-2', '5-3-2', '3-4-3', '4-1-4-1']


def _save_upload(file, subfolder):
    """Save uploaded image, return relative URL or None."""
    from austinapp import allowed_file
    if not file or not file.filename:
        return None
    if not allowed_file(file.filename):
        return None
    ext = file.filename.rsplit('.', 1)[1].lower()
    fname = f'{subfolder}_{uuid.uuid4().hex[:10]}.{ext}'
    path = os.path.join(current_app.config['UPLOAD_FOLDER'], subfolder, fname)
    file.save(path)
    return f'uploads/{subfolder}/{fname}'


# ─── List ─────────────────────────────────────────────────────

@team_bp.route('/teams')
@login_required
def list_teams():
    search = request.args.get('q', '').strip()
    sport_filter = request.args.get('sport', '')
    query = Team.query
    if search:
        query = query.filter(Team.name.ilike(f'%{search}%'))
    if sport_filter:
        query = query.filter_by(sport=sport_filter)
    teams = query.order_by(Team.name).all()
    return render_template('team/list.html', teams=teams, search=search, sport_filter=sport_filter)


# ─── Create ───────────────────────────────────────────────────

@team_bp.route('/teams/create', methods=['GET', 'POST'])
@login_required
@require_manager
def create_team():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('O nome é obrigatório.', 'error')
            return render_template('team/create.html', formations_football=FORMATIONS_FOOTBALL, formations_futsal=FORMATIONS_FUTSAL)
        if Team.query.filter_by(name=name).first():
            flash('Já existe uma equipa com esse nome.', 'error')
            return render_template('team/create.html', formations_football=FORMATIONS_FOOTBALL, formations_futsal=FORMATIONS_FUTSAL)

        team = Team(
            name=name,
            short_name=request.form.get('short_name', '').strip().upper(),
            founded_date=request.form.get('founded_date', '').strip(),
            city=request.form.get('city', '').strip(),
            country=request.form.get('country', 'Portugal').strip(),
            primary_color=request.form.get('primary_color', '#2563eb'),
            secondary_color=request.form.get('secondary_color', '#ffffff'),
            stadium=request.form.get('stadium', '').strip(),
            ceo=request.form.get('ceo', '').strip(),
            manager=request.form.get('manager', '').strip(),
            coach=request.form.get('coach', '').strip(),
            assistant_coach=request.form.get('assistant_coach', '').strip(),
            captain=request.form.get('captain', '').strip(),
            formation=request.form.get('formation', '4-3-3'),
            tactic=request.form.get('tactic', '').strip(),
            sport=request.form.get('sport', 'futsal'),
            email=request.form.get('email', '').strip(),
            phone=request.form.get('phone', '').strip(),
            website=request.form.get('website', '').strip(),
            instagram=request.form.get('instagram', '').strip(),
            notes=request.form.get('notes', '').strip(),
        )
        db.session.add(team)
        db.session.commit()

        # Logo upload
        logo_url = _save_upload(request.files.get('logo'), 'logos')
        if logo_url:
            team.logo_url = logo_url
            db.session.commit()

        flash(f'Equipa "{name}" criada! ⚽', 'success')
        return redirect(url_for('team.team_room', team_id=team.id))

    return render_template('team/create.html', formations_football=FORMATIONS_FOOTBALL, formations_futsal=FORMATIONS_FUTSAL)


# ─── Room ─────────────────────────────────────────────────────

@team_bp.route('/teams/<int:team_id>')
@login_required
def team_room(team_id):
    team = Team.query.get_or_404(team_id)
    players = Player.query.filter_by(team_id=team_id).order_by(Player.number).all()
    posts = TeamPost.query.filter_by(team_id=team_id).order_by(TeamPost.created_at.desc()).limit(20).all()

    active = [p for p in players if p.status == 'ativo']
    injured = [p for p in players if p.status == 'lesionado']
    suspended = [p for p in players if p.status == 'suspenso']
    inactive = [p for p in players if p.status == 'inativo']

    # Next matches for this team
    from models.match import Match
    from datetime import datetime
    today = datetime.today().strftime('%Y-%m-%d')
    next_matches = Match.query.filter(
        db.or_(Match.team1_id == team_id, Match.team2_id == team_id),
        Match.status == 'scheduled',
        Match.date >= today
    ).order_by(Match.date).limit(5).all()

    # Feed da equipa
    from utils.feed_helpers import get_team_feed
    team_feed_posts = get_team_feed(team_id, limit=20)

    return render_template('team/room.html', team=team, players=players, posts=posts,
                           active=active, injured=injured, suspended=suspended, inactive=inactive,
                           positions=ALL_POSITIONS, next_matches=next_matches,
                           team_feed_posts=team_feed_posts)


# ─── Edit ─────────────────────────────────────────────────────

@team_bp.route('/teams/<int:team_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_team(team_id):
    team = Team.query.get_or_404(team_id)
    if request.method == 'POST':
        team.name = request.form.get('name', team.name).strip()
        team.short_name = request.form.get('short_name', team.short_name or '').strip().upper()
        team.founded_date = request.form.get('founded_date', team.founded_date or '').strip()
        team.city = request.form.get('city', team.city or '').strip()
        team.country = request.form.get('country', team.country or 'Portugal').strip()
        team.primary_color = request.form.get('primary_color', team.primary_color or '#2563eb')
        team.secondary_color = request.form.get('secondary_color', team.secondary_color or '#ffffff')
        team.stadium = request.form.get('stadium', team.stadium or '').strip()
        team.ceo = request.form.get('ceo', team.ceo or '').strip()
        team.manager = request.form.get('manager', team.manager or '').strip()
        team.coach = request.form.get('coach', team.coach or '').strip()
        team.assistant_coach = request.form.get('assistant_coach', team.assistant_coach or '').strip()
        team.captain = request.form.get('captain', team.captain or '').strip()
        team.formation = request.form.get('formation', team.formation)
        team.tactic = request.form.get('tactic', team.tactic or '').strip()
        team.sport = request.form.get('sport', team.sport or 'futsal')
        team.email = request.form.get('email', team.email or '').strip()
        team.phone = request.form.get('phone', team.phone or '').strip()
        team.website = request.form.get('website', team.website or '').strip()
        team.instagram = request.form.get('instagram', team.instagram or '').strip()
        team.notes = request.form.get('notes', team.notes or '').strip()

        logo_url = _save_upload(request.files.get('logo'), 'logos')
        if logo_url:
            team.logo_url = logo_url

        db.session.commit()
        flash('Equipa atualizada! ✅', 'success')
        return redirect(url_for('team.team_room', team_id=team.id))

    return render_template('team/edit.html', team=team,
                           formations_football=FORMATIONS_FOOTBALL, formations_futsal=FORMATIONS_FUTSAL)


# ─── Delete ───────────────────────────────────────────────────

@team_bp.route('/teams/<int:team_id>/delete', methods=['POST'])
@login_required
def delete_team(team_id):
    team = Team.query.get_or_404(team_id)
    db.session.delete(team)
    db.session.commit()
    flash('Equipa eliminada.', 'info')
    return redirect(url_for('team.list_teams'))


# ─── Add Player ───────────────────────────────────────────────

@team_bp.route('/teams/<int:team_id>/players/add', methods=['POST'])
@login_required
def add_player(team_id):
    Team.query.get_or_404(team_id)
    player = Player(
        name=request.form.get('name', '').strip(),
        nickname=request.form.get('nickname', '').strip(),
        number=request.form.get('number', type=int),
        position=request.form.get('position', '').strip(),
        secondary_position=request.form.get('secondary_position', '').strip(),
        birth_date=request.form.get('birth_date', '').strip(),
        age=request.form.get('age', type=int),
        nationality=request.form.get('nationality', '').strip(),
        id_number=request.form.get('id_number', '').strip(),
        phone=request.form.get('phone', '').strip(),
        email=request.form.get('email', '').strip(),
        height_cm=request.form.get('height_cm', type=int),
        weight_kg=request.form.get('weight_kg', type=int),
        dominant_foot=request.form.get('dominant_foot', 'Direito'),
        status=request.form.get('status', 'ativo'),
        contract_start=request.form.get('contract_start', '').strip(),
        contract_end=request.form.get('contract_end', '').strip(),
        team_id=team_id,
    )
    db.session.add(player)
    db.session.commit()

    photo_url = _save_upload(request.files.get('photo'), 'players')
    if photo_url:
        player.photo_url = photo_url
        db.session.commit()

    flash(f'Jogador {player.name} adicionado! ✅', 'success')
    return redirect(url_for('team.team_room', team_id=team_id))


# ─── Edit Player ──────────────────────────────────────────────

@team_bp.route('/players/<int:player_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_player(player_id):
    player = Player.query.get_or_404(player_id)
    if request.method == 'POST':
        player.name = request.form.get('name', player.name).strip()
        player.nickname = request.form.get('nickname', player.nickname or '').strip()
        player.number = request.form.get('number', type=int) or player.number
        player.position = request.form.get('position', player.position or '').strip()
        player.secondary_position = request.form.get('secondary_position', player.secondary_position or '').strip()
        player.birth_date = request.form.get('birth_date', player.birth_date or '').strip()
        player.age = request.form.get('age', type=int) or player.age
        player.nationality = request.form.get('nationality', player.nationality or '').strip()
        player.id_number = request.form.get('id_number', player.id_number or '').strip()
        player.phone = request.form.get('phone', player.phone or '').strip()
        player.email = request.form.get('email', player.email or '').strip()
        player.height_cm = request.form.get('height_cm', type=int) or player.height_cm
        player.weight_kg = request.form.get('weight_kg', type=int) or player.weight_kg
        player.dominant_foot = request.form.get('dominant_foot', player.dominant_foot or 'Direito')
        player.status = request.form.get('status', player.status or 'ativo')
        player.contract_start = request.form.get('contract_start', player.contract_start or '').strip()
        player.contract_end = request.form.get('contract_end', player.contract_end or '').strip()

        photo_url = _save_upload(request.files.get('photo'), 'players')
        if photo_url:
            player.photo_url = photo_url

        db.session.commit()
        flash('Jogador atualizado! ✅', 'success')
        return redirect(url_for('team.team_room', team_id=player.team_id))

    return render_template('team/edit_player.html', player=player, positions=ALL_POSITIONS)


# ─── Delete Player ────────────────────────────────────────────

@team_bp.route('/players/<int:player_id>/delete', methods=['POST'])
@login_required
def delete_player(player_id):
    player = Player.query.get_or_404(player_id)
    team_id = player.team_id
    db.session.delete(player)
    db.session.commit()
    flash('Jogador removido.', 'info')
    return redirect(url_for('team.team_room', team_id=team_id))


# ─── Team Posts ───────────────────────────────────────────────

@team_bp.route('/teams/<int:team_id>/posts/add', methods=['POST'])
@login_required
def add_post(team_id):
    Team.query.get_or_404(team_id)
    content = request.form.get('content', '').strip()
    post_type = request.form.get('post_type', 'update')

    if not content:
        flash('Escreve algo antes de publicar.', 'error')
        return redirect(url_for('team.team_room', team_id=team_id))

    post = TeamPost(
        team_id=team_id,
        author_id=current_user.id,
        content=content,
        post_type=post_type,
    )
    db.session.add(post)
    db.session.commit()

    image_url = _save_upload(request.files.get('image'), 'posts')
    if image_url:
        post.image_url = image_url
        db.session.commit()

    flash('Publicação adicionada! 📢', 'success')
    return redirect(url_for('team.team_room', team_id=team_id))


@team_bp.route('/teams/posts/<int:post_id>/delete', methods=['POST'])
@login_required
def delete_post(post_id):
    post = TeamPost.query.get_or_404(post_id)
    team_id = post.team_id
    db.session.delete(post)
    db.session.commit()
    flash('Publicação removida.', 'info')
    return redirect(url_for('team.team_room', team_id=team_id))


# ─── Excel Import / Export ────────────────────────────────────

@team_bp.route('/teams/<int:team_id>/players/import', methods=['POST'])
@login_required
def import_players_excel(team_id):
    Team.query.get_or_404(team_id)
    if 'excel_file' not in request.files:
        flash('Nenhum ficheiro selecionado.', 'error')
        return redirect(url_for('team.team_room', team_id=team_id))

    file = request.files['excel_file']
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Formato inválido. Usa .xlsx ou .xls', 'error')
        return redirect(url_for('team.team_room', team_id=team_id))

    try:
        # ── Carregar workbook ─────────────────────────────────
        # read_only=False para compatibilidade máxima com ficheiros .xls/.xlsx
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # ── Ler headers da primeira linha ─────────────────────
        # Normalizar: minúsculas, sem espaços extra, sem acentos problemáticos
        def norm(s):
            """Normaliza header: minúsculas, strip, remove acentos comuns."""
            if s is None:
                return ''
            s = str(s).strip().lower()
            # Substituições de acentos para matching robusto
            replacements = {
                'ã': 'a', 'á': 'a', 'à': 'a', 'â': 'a',
                'é': 'e', 'ê': 'e', 'è': 'e',
                'í': 'i', 'î': 'i',
                'ó': 'o', 'ô': 'o', 'õ': 'o',
                'ú': 'u', 'û': 'u',
                'ç': 'c',
                'º': 'o', 'ª': 'a',
            }
            for orig, rep in replacements.items():
                s = s.replace(orig, rep)
            return s

        # Ler todas as linhas como lista de listas
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            flash('Ficheiro Excel vazio.', 'error')
            return redirect(url_for('team.team_room', team_id=team_id))

        # Encontrar a linha de headers (primeira linha não vazia)
        header_row_idx = 0
        for i, row in enumerate(all_rows):
            if any(v is not None for v in row):
                header_row_idx = i
                break

        raw_headers = all_rows[header_row_idx]
        headers_norm = [norm(h) for h in raw_headers]

        # Mapa de header normalizado → índice da coluna
        col = {h: i for i, h in enumerate(headers_norm) if h}

        # ── Função para obter valor de uma célula ─────────────
        def get_val(row, *keys):
            """Tenta vários nomes de coluna, retorna string ou ''."""
            for k in keys:
                k_norm = norm(k)
                idx = col.get(k_norm)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    val = str(row[idx]).strip()
                    if val and val.lower() not in ('none', 'nan', '-', '—'):
                        return val
            return ''

        def get_int_val(row, *keys):
            """Tenta vários nomes de coluna, retorna int ou None."""
            for k in keys:
                k_norm = norm(k)
                idx = col.get(k_norm)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    try:
                        # Pode vir como float (ex: 25.0) do Excel
                        return int(float(str(row[idx]).strip()))
                    except (TypeError, ValueError):
                        pass
            return None

        # ── Processar linhas de dados ─────────────────────────
        imported = updated = skipped = errors = 0
        error_details = []

        data_rows = all_rows[header_row_idx + 1:]

        for row_num, row in enumerate(data_rows, start=header_row_idx + 2):
            # Ignorar linhas completamente vazias
            if all(v is None or str(v).strip() == '' for v in row):
                continue

            try:
                # Nome é o único campo obrigatório
                name = get_val(row,
                    'nome', 'name', 'nome completo', 'full name',
                    'jogador', 'player', 'nome do jogador')

                if not name:
                    # Tentar usar a primeira coluna não vazia como nome
                    for v in row:
                        if v is not None and str(v).strip():
                            name = str(v).strip()
                            break

                if not name:
                    skipped += 1
                    continue

                # Limpar nome (remover números de linha acidentais)
                if name.isdigit():
                    skipped += 1
                    continue

                # Recolher todos os campos (incompletos são aceites)
                nickname      = get_val(row, 'alcunha', 'nickname', 'apelido', 'nome guerra')
                number_raw    = get_int_val(row, 'no', 'numero', 'number', 'num', 'camisa', 'camisola')
                position      = get_val(row, 'posicao', 'posição', 'position', 'pos', 'funcao')
                sec_position  = get_val(row, 'posicao secundaria', 'secondary position', 'pos2')
                birth_date    = get_val(row, 'data nascimento', 'nascimento', 'birth date',
                                        'birth_date', 'data de nascimento', 'dob')
                age_raw       = get_int_val(row, 'idade', 'age', 'anos')
                nationality   = get_val(row, 'nacionalidade', 'nationality', 'pais', 'país')
                id_number     = get_val(row, 'cc', 'id', 'passaporte', 'bi', 'documento',
                                        'id number', 'nif', 'bilhete')
                phone         = get_val(row, 'telefone', 'phone', 'telemovel', 'tel', 'contacto')
                email         = get_val(row, 'email', 'e-mail', 'mail')
                height_raw    = get_int_val(row, 'altura', 'height', 'alt', 'height cm')
                weight_raw    = get_int_val(row, 'peso', 'weight', 'kg', 'weight kg')
                dominant_foot = get_val(row, 'pe', 'pé', 'foot', 'pe dominante', 'dominant foot') or 'Direito'
                status_raw    = get_val(row, 'estado', 'status', 'situacao', 'situação') or 'ativo'
                contract_start = get_val(row, 'contrato inicio', 'contract start',
                                         'contract_start', 'inicio contrato', 'data inicio')
                contract_end   = get_val(row, 'contrato fim', 'contract end',
                                         'contract_end', 'fim contrato', 'data fim', 'validade')

                # Normalizar status
                status_map = {
                    'ativo': 'ativo', 'active': 'ativo', 'activo': 'ativo',
                    'lesionado': 'lesionado', 'injured': 'lesionado', 'lesao': 'lesionado',
                    'suspenso': 'suspenso', 'suspended': 'suspenso',
                    'inativo': 'inativo', 'inactive': 'inativo', 'inactivo': 'inativo',
                    'pendente': 'pendente', 'pending': 'pendente',
                }
                status = status_map.get(norm(status_raw), 'ativo')

                # Normalizar pé dominante
                foot_map = {
                    'direito': 'Direito', 'right': 'Direito', 'd': 'Direito',
                    'esquerdo': 'Esquerdo', 'left': 'Esquerdo', 'e': 'Esquerdo',
                    'ambos': 'Ambos', 'both': 'Ambos',
                }
                dominant_foot = foot_map.get(norm(dominant_foot), dominant_foot or 'Direito')

                # ── Verificar se jogador já existe (pelo nome + equipa) ──
                existing = Player.query.filter_by(name=name, team_id=team_id).first()

                if existing:
                    # Atualizar campos que vieram preenchidos (não sobrescrever com vazio)
                    if nickname:        existing.nickname       = nickname
                    if number_raw:      existing.number         = number_raw
                    if position:        existing.position       = position
                    if sec_position:    existing.secondary_position = sec_position
                    if birth_date:      existing.birth_date     = birth_date
                    if age_raw:         existing.age            = age_raw
                    if nationality:     existing.nationality    = nationality
                    if id_number:       existing.id_number      = id_number
                    if phone:           existing.phone          = phone
                    if email:           existing.email          = email
                    if height_raw:      existing.height_cm      = height_raw
                    if weight_raw:      existing.weight_kg      = weight_raw
                    existing.dominant_foot  = dominant_foot
                    existing.status         = status
                    if contract_start:  existing.contract_start = contract_start
                    if contract_end:    existing.contract_end   = contract_end
                    updated += 1
                else:
                    # Criar novo jogador (campos incompletos são aceites)
                    player = Player(
                        name=name,
                        nickname=nickname or None,
                        number=number_raw,
                        position=position or None,
                        secondary_position=sec_position or None,
                        birth_date=birth_date or None,
                        age=age_raw,
                        nationality=nationality or None,
                        id_number=id_number or None,
                        phone=phone or None,
                        email=email or None,
                        height_cm=height_raw,
                        weight_kg=weight_raw,
                        dominant_foot=dominant_foot,
                        status=status,
                        contract_start=contract_start or None,
                        contract_end=contract_end or None,
                        team_id=team_id,
                    )
                    db.session.add(player)
                    imported += 1

            except Exception as row_err:
                errors += 1
                error_details.append(f'Linha {row_num}: {row_err}')
                # Continuar para a próxima linha em vez de abortar tudo
                db.session.rollback()
                continue

        # ── Commit final ──────────────────────────────────────
        try:
            db.session.commit()
        except Exception as commit_err:
            db.session.rollback()
            flash(f'❌ Erro ao guardar: {commit_err}', 'error')
            return redirect(url_for('team.team_room', team_id=team_id))

        wb.close()

        # ── Mensagem de resultado ─────────────────────────────
        parts = []
        if imported:  parts.append(f'✅ {imported} importados')
        if updated:   parts.append(f'🔄 {updated} atualizados')
        if skipped:   parts.append(f'⏭️ {skipped} ignorados (sem nome)')
        if errors:    parts.append(f'⚠️ {errors} com erro')

        msg = ' · '.join(parts) if parts else 'Nenhum jogador processado.'
        flash(msg, 'success' if (imported or updated) else 'info')

        # Mostrar detalhes de erros se houver
        for detail in error_details[:5]:  # máximo 5 erros detalhados
            flash(detail, 'error')

    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao ler o ficheiro Excel: {e}', 'error')

    return redirect(url_for('team.team_room', team_id=team_id))


@team_bp.route('/teams/excel-template')
@login_required
def download_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Jogadores'

    # Estilos
    hf    = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    ex_fill = PatternFill(start_color='0D1526', end_color='0D1526', fill_type='solid')
    ex_font = Font(color='4B6080', italic=True, size=10)

    # Colunas: (header, largura, exemplo)
    columns = [
        ('Nome *',           25, 'Cristiano Ronaldo'),
        ('Alcunha',          18, 'CR7'),
        ('Nº',                6, '7'),
        ('Posição',          12, 'ALA-D'),
        ('Data Nascimento',  18, '1985-02-05'),
        ('Idade',             8, '39'),
        ('Nacionalidade',    18, 'Portuguesa'),
        ('CC',               16, '12345678'),
        ('Telefone',         16, '+351 912 345 678'),
        ('Email',            25, 'cr7@austin.com'),
        ('Altura',           10, '187'),
        ('Peso',              8, '83'),
        ('Pé',               10, 'Direito'),
        ('Estado',           12, 'ativo'),
        ('Contrato Inicio',  16, '2024-01-01'),
        ('Contrato Fim',     16, '2025-12-31'),
    ]

    # Linha 1 — Headers
    for i, (h, w, _) in enumerate(columns, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hf
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = w

    # Linha 2 — Exemplo (em cinzento)
    for i, (_, _, ex) in enumerate(columns, 1):
        c = ws.cell(row=2, column=i, value=ex)
        c.font = ex_font
        c.fill = ex_fill
        c.alignment = Alignment(horizontal='center')

    # Linha 3 — Linha vazia para o utilizador começar a preencher
    # (deixar em branco)

    # Nota na célula A4
    ws.cell(row=4, column=1,
            value='⚠️ A linha 2 é apenas um exemplo — podes apagá-la. Só o campo "Nome *" é obrigatório.')
    ws.cell(row=4, column=1).font = Font(color='F59E0B', italic=True, size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='ALC_template_jogadores.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@team_bp.route('/teams/<int:team_id>/lineup/save', methods=['POST'])
@login_required
def save_team_lineup(team_id):
    """Save the tactical lineup for a team (not match-specific)."""
    from flask import jsonify
    team = Team.query.get_or_404(team_id)
    data = request.get_json()

    if not data:
        return jsonify({'ok': False, 'error': 'No data'}), 400

    formation = data.get('formation', team.formation)
    starters = data.get('starters', [])  # [{slot, player_id}]

    # Update team formation
    team.formation = formation

    # Store lineup as captain field hack — actually store in notes as JSON
    # In a real system you'd have a TeamLineup model; here we use a simple approach
    import json
    lineup_json = json.dumps({
        'formation': formation,
        'starters': starters
    })
    team.notes = (team.notes or '').split('__LINEUP__')[0] + f'__LINEUP__{lineup_json}'
    db.session.commit()

    return jsonify({'ok': True, 'formation': formation, 'count': len(starters)})


@team_bp.route('/teams/<int:team_id>/export')
@login_required
def export_team_excel(team_id):
    team = Team.query.get_or_404(team_id)
    players = Player.query.filter_by(team_id=team_id).order_by(Player.number).all()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = team.name[:30]
    ws['A1'] = f'AUSTIN LEAGUE CORE — {team.name}'
    ws['A1'].font = Font(bold=True, size=14, color='2563EB')
    hf = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    headers = ['Nº','Nome','Alcunha','Posição','Idade','Nac.','Alt.','Peso','Pé','Estado','⚽','🎯','🟨','🟥','Jogos']
    for i,h in enumerate(headers,1):
        c = ws.cell(row=3,column=i,value=h)
        c.font=Font(bold=True,color='FFFFFF'); c.fill=hf
    for ri,p in enumerate(players,4):
        row=[p.number,p.name,p.nickname,p.position,p.age,p.nationality,
             p.height_cm,p.weight_kg,p.dominant_foot,p.status,
             p.goals,p.assists,p.yellow_cards,p.red_cards,p.matches_played]
        for ci,v in enumerate(row,1):
            ws.cell(row=ri,column=ci,value=v)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'ALC_{team.name.replace(" ","_")}_plantel.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
